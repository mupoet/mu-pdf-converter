#!/usr/bin/env python3
"""
pdf_to_xlsx.py - 提取 PDF 中的表格并写入 Excel
用法:
  单文件:python pdf_to_xlsx.py input.pdf [--pages all|1-3] [--outfile output.xlsx]
  批量:  python pdf_to_xlsx.py --batch ./invoices/ [--outfile batch_result.xlsx]
"""

import argparse
import sys
from pathlib import Path

# ── 依赖检查 ──────────────────────────────────
_missing = []
try:
    import pdfplumber
except ImportError:
    _missing.append("pdfplumber")
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    _missing.append("openpyxl")

try:
    from markitdown.converters._pdf_converter import _extract_form_content_from_words as _markitdown_extract
    _HAS_MARKITDOWN = True
except ImportError:
    _HAS_MARKITDOWN = False

try:
    import pypdfium2 as pdfium
    _HAS_PDFIUM = True
except ImportError:
    _HAS_PDFIUM = False

if _missing:
    print(
        f"[错误] 缺少依赖:{', '.join(_missing)}\n"
        f"请运行:pip install {' '.join(_missing)}",
        file=sys.stderr,
    )
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from utils import parse_page_range, generate_output_path, xy_cut_extract_tables, parse_markdown_tables, check_scanned_and_warn  # noqa: E402


# ─────────────────────────────────────────────
# 引擎层:三级表格提取优先级
#   E1: pypdfium2 + XY-Cut (无边框识别,精u最高)
#   E2: MarkItDown  (AI辅助,处理复杂排版)
#   E3: pdfplumber  (传统线段检测,有边框表格最稳)
# ─────────────────────────────────────────────

# _parse_markdown_tables 已移至 utils.py 公共模块（parse_markdown_tables）


# ─────────────────────────────────────────────
# 核心逻辑
# ─────────────────────────────────────────────

def _extract_tables_pdfium_xycut(pdf_path: str, page_idx: int) -> list:
    """
    E1 引擎:用 pypdfium2 提取页面词语带坐标,再经 XY-Cut 识别表格。
    返回二维 list(表格组成的 rows)的列表。
    """
    if not _HAS_PDFIUM:
        return []
    try:
        doc = pdfium.PdfDocument(pdf_path)
        page = doc[page_idx]
        width = page.get_width()
        height = page.get_height()
        textpage = page.get_textpage()

        # 提取每个词的坐标和文本
        words = []
        n_chars = textpage.count_chars()
        if n_chars == 0:
            doc.close()
            return []

        # 用 get_text_bounded 提取分段块,再拆词
        # 更可靠:用 search 或 char-level bbox 聚合成行
        # 这里用 pdfplumber words 格式引入 pypdfium2 的 char级 bbox
        i = 0
        while i < n_chars:
            ch = textpage.get_charbox(i, loose=True)
            char_text = textpage.get_text_range(i, 1)
            if char_text.strip():
                words.append({
                    'x0': ch[0], 'top': height - ch[3],
                    'x1': ch[2], 'bottom': height - ch[1],
                    'text': char_text,
                })
            i += 1

        doc.close()

        # 先把字符聚合成词语(相邻字符 x 间距 < word_gap)
        word_gap = width * 0.012
        merged = _merge_chars_to_words(words, word_gap)

        return xy_cut_extract_tables(merged, width, height)
    except Exception as e:
        print(f"    [E1引擎] pypdfium2+XY-Cut 失败({e}),降级", file=sys.stderr)
        return []


def _merge_chars_to_words(chars: list, word_gap: float) -> list:
    """
    将字符级指幵聚合成词语,相邻字符同行且 x 间距小于 word_gap 则合并。
    返回合并后的 words 列表。
    """
    if not chars:
        return []
    # 按行组(top 相近)
    row_gap = (max(c['bottom'] for c in chars) - min(c['top'] for c in chars)) * 0.015
    lines = {}
    for c in chars:
        row_key = round(c['top'] / max(row_gap, 1))
        lines.setdefault(row_key, []).append(c)

    words = []
    for _, line_chars in sorted(lines.items()):
        line_chars.sort(key=lambda c: c['x0'])
        cur = line_chars[0].copy()
        for ch in line_chars[1:]:
            if ch['x0'] - cur['x1'] <= word_gap:
                cur['text'] += ch['text']
                cur['x1'] = max(cur['x1'], ch['x1'])
                cur['bottom'] = max(cur['bottom'], ch['bottom'])
            else:
                if cur['text'].strip():
                    words.append(cur)
                cur = ch.copy()
        if cur['text'].strip():
            words.append(cur)
    return words


def pdf_to_xlsx(
    pdf_path: str,
    pages_str: str = "all",
    outfile: str = None,
) -> str:
    """
    提取 PDF 所有页面中的表格,写入 Excel 文件。

    表格提取优先级:
      E1 pypdfium2+XY-Cut → E2 MarkItDown → E3 pdfplumber
    每个表格写入独立 Sheet,Sheet 名:Page{n}_Table{m}。

    参数:
        pdf_path: 输入 PDF 路径
        pages_str: 页面范围
        outfile: 输出 Excel 路径(可选)
    返回:
        输出文件路径
    """
    pdf_path = str(Path(pdf_path).resolve())
    if not Path(pdf_path).exists():
        print(f"[错误] 文件不存在:{pdf_path}", file=sys.stderr)
        sys.exit(1)

    output_path = generate_output_path(pdf_path, ".xlsx", outfile=outfile)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet

    total_tables = 0

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        page_indices = parse_page_range(pages_str, total_pages)
        print(f"[信息] 共 {total_pages} 页,扫描范围:{len(page_indices)} 页")

        for idx in page_indices:
            page = pdf.pages[idx]
            page_num = idx + 1

            # E1: pypdfium2 + XY-Cut(无边框表格首选)
            tables = None
            if _HAS_PDFIUM:
                e1_tables = _extract_tables_pdfium_xycut(pdf_path, idx)
                if e1_tables:
                    tables = e1_tables
                    print(f"  - 第 {page_num} 页:[E1 XY-Cut] 识别到 {len(tables)} 个表格")

            # E2: MarkItDown(E1 无结果时尝试)
            if tables is None and _HAS_MARKITDOWN:
                try:
                    md_result = _markitdown_extract(page)
                    if md_result is not None:
                        tables = parse_markdown_tables(md_result)
                        if tables:
                            print(f"  - 第 {page_num} 页:[E2 MarkItDown] 识别到 {len(tables)} 个表格")
                        else:
                            tables = None
                except Exception as e:
                    print(f"  - 第 {page_num} 页:[E2 MarkItDown] 失败({e}),继续降级")
                    tables = None

            # E3: pdfplumber(最终化递,有边框表格最可靠)
            if tables is None:
                tables = page.extract_tables()
                if tables:
                    print(f"  - 第 {page_num} 页:[E3 pdfplumber] 识别到 {len(tables)} 个表格")

            if not tables:
                print(f"  - 第 {page_num} 页:未发现表格")
                page.close()
                continue

            for t_idx, table in enumerate(tables):
                t_num = t_idx + 1
                sheet_name = f"Page{page_num}_Table{t_num}"
                # Sheet 名最长 31 字符(Excel 限制)
                sheet_name = sheet_name[:31]
                ws = wb.create_sheet(title=sheet_name)

                if not table:
                    print(f"  - 第 {page_num} 页第 {t_num} 个表格:为空,跳过")
                    continue

                # 写入数据
                for row_idx, row in enumerate(table):
                    for col_idx, cell_val in enumerate(row):
                        cell_val = cell_val or ""
                        # 清理换行符
                        cell_val = str(cell_val).replace("\n", " ").strip()
                        cell = ws.cell(
                            row=row_idx + 1,
                            column=col_idx + 1,
                            value=cell_val,
                        )
                        # 首行加粗 + 淡蓝色背景
                        if row_idx == 0:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color="D9EAF7",
                                end_color="D9EAF7",
                                fill_type="solid",
                            )
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

                # 列宽自适应
                _auto_column_width(ws)

                total_tables += 1
                print(f"  ✓ 第 {page_num} 页,第 {t_num} 个表格 → Sheet: {sheet_name}")

            page.close()  # 修复内存泄漏

    if total_tables == 0:
        print("\n[提示] 未在 PDF 中发现任何表格。")
        print("       可能原因:")
        print("       1. PDF 中的表格为图片形式(扫描件),无法通过文本提取")
        print("       2. 表格没有明显的边框线,pdfplumber 无法识别")
        print("       3. 该 PDF 确实不含表格")
        print("\n       建议:使用 pdf_to_images.py 将 PDF 转为图片后手动查看")
        return None

    wb.save(output_path)
    print(f"\n[完成] 共提取 {total_tables} 个表格 → {output_path}")
    return output_path


# ─────────────────────────────────────────────
# 批量模式:扫描文件夹内所有 PDF
# ─────────────────────────────────────────────

def _extract_tables_from_pdf(pdf_path: str, pages_str: str = "all") -> list:
    """
    提取单个 PDF 中所有表格，返回列表：
    [{"page": int, "table_idx": int, "rows": [[...]], "file": str}]
    引擎优先级： E1 XY-Cut → E2 MarkItDown → E3 pdfplumber
    """
    results = []
    file_stem = Path(pdf_path).stem

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        page_indices = parse_page_range(pages_str, total_pages)

        for idx in page_indices:
            page = pdf.pages[idx]
            page_num = idx + 1

            # E1: pypdfium2 + XY-Cut
            tables = None
            if _HAS_PDFIUM:
                e1_tables = _extract_tables_pdfium_xycut(pdf_path, idx)
                if e1_tables:
                    tables = e1_tables

            # E2: MarkItDown
            if tables is None and _HAS_MARKITDOWN:
                try:
                    md_result = _markitdown_extract(page)
                    if md_result is not None:
                        tables = parse_markdown_tables(md_result)
                        if not tables:
                            tables = None
                except Exception:
                    tables = None

            # E3: pdfplumber
            if tables is None:
                tables = page.extract_tables()

            if not tables:
                page.close()
                continue

            for t_idx, table in enumerate(tables):
                results.append({
                    "file": file_stem,
                    "page": page_num,
                    "table_idx": t_idx + 1,
                    "rows": table or [],
                })

            page.close()

    return results


def batch_pdf_to_xlsx(folder: str, outfile: str = None) -> str:
    """
    扫描 folder 内所有 .pdf 文件,提取所有表格写入同一个 Excel。

    Sheet 命名规则:{文件名}_{P页}_{T表}(截断到 31 字符)
    最后一个 Sheet 为汇总 Summary。

    参数:
        folder: 包含 PDF 文件的目录
        outfile: 输出 Excel 路径(默认 ./pdf_batch_tables.xlsx)
    返回:
        输出文件路径
    """
    folder_path = Path(folder).resolve()
    if not folder_path.is_dir():
        print(f"[错误] 目录不存在:{folder}", file=sys.stderr)
        sys.exit(1)

    MAX_BATCH_FILES = 100  # 单次批量处理上限，防止内存溢出

    pdf_files = sorted(folder_path.glob("*.pdf"))
    if not pdf_files:
        print(f"[提示] 目录中未找到 .pdf 文件:{folder}", file=sys.stderr)
        sys.exit(1)

    if len(pdf_files) > MAX_BATCH_FILES:
        print(f"[警告] 目录含 {len(pdf_files)} 个 PDF，超过单次上限 {MAX_BATCH_FILES}，仅处理前 {MAX_BATCH_FILES} 个",
              file=sys.stderr)
        pdf_files = pdf_files[:MAX_BATCH_FILES]

    print(f"[批量] 扫描目录:{folder_path}")
    print(f"[批量] 发现 {len(pdf_files)} 个 PDF 文件")

    out_path = outfile or "./pdf_batch_tables.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet

    summary_rows = []  # 汇总数据
    total_tables = 0

    for pdf_file in pdf_files:
        print(f"\n  处理:{pdf_file.name}")
        try:
            tables = _extract_tables_from_pdf(str(pdf_file))
        except Exception as e:
            print(f"  [警告] 处理失败:{e}", file=sys.stderr)
            continue

        if not tables:
            print(f"  - 未发现表格")
            continue

        for tbl in tables:
            file_stem = tbl["file"]
            page_num = tbl["page"]
            t_idx = tbl["table_idx"]
            rows = tbl["rows"]

            # Sheet 名:{文件名}_{P页}_{T表},最长 31 字符
            sheet_name = f"{file_stem}_P{page_num}_T{t_idx}"
            sheet_name = sheet_name[:31]

            # 避免 Sheet 名重复
            existing = {ws.title for ws in wb.worksheets}
            base_name = sheet_name
            counter = 2
            while sheet_name in existing:
                sheet_name = f"{base_name[:28]}_{counter}"
                counter += 1

            ws = wb.create_sheet(title=sheet_name)

            row_count = 0
            col_count = 0
            for r_idx, row in enumerate(rows):
                if not row:
                    continue
                col_count = max(col_count, len(row))
                for c_idx, cell_val in enumerate(row):
                    cell_val = str(cell_val or "").replace("\n", " ").strip()
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=cell_val)
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="D9EAF7",
                            end_color="D9EAF7",
                            fill_type="solid",
                        )
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row_count += 1

            _auto_column_width(ws)
            total_tables += 1

            summary_rows.append({
                "file": pdf_file.name,
                "page": page_num,
                "table_idx": t_idx,
                "rows": row_count,
                "cols": col_count,
                "sheet": sheet_name,
            })
            print(f"    ✓ P{page_num}_T{t_idx} → Sheet: {sheet_name}  ({row_count}行 × {col_count}列)")

    if total_tables == 0:
        print("\n[提示] 所有 PDF 中均未发现表格。")
        return None

    # 生成汇总 Sheet
    ws_summary = wb.create_sheet(title="Summary")
    headers = ["文件名", "页码", "表序号", "行数", "列数", "Sheet 名"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws_summary.cell(row=1, column=c_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(summary_rows, start=2):
        ws_summary.cell(row=r_idx, column=1, value=row["file"])
        ws_summary.cell(row=r_idx, column=2, value=row["page"])
        ws_summary.cell(row=r_idx, column=3, value=row["table_idx"])
        ws_summary.cell(row=r_idx, column=4, value=row["rows"])
        ws_summary.cell(row=r_idx, column=5, value=row["cols"])
        ws_summary.cell(row=r_idx, column=6, value=row["sheet"])

    _auto_column_width(ws_summary)

    wb.save(out_path)
    print(f"\n[完成] 批量提取完成:{len(pdf_files)} 个 PDF,{total_tables} 个表格 → {out_path}")
    return out_path


def _auto_column_width(ws, max_width: int = 50, min_width: int = 8):
    """自动设置列宽(基于内容最大字符数)"""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # 中文字符宽度约为英文 2 倍
                val_str = str(cell.value)
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ─────────────────────────────────────────────
# CLI 入口
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="提取 PDF 中的表格并写入 Excel(每个表格一个 Sheet)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例(单文件):
  python pdf_to_xlsx.py report.pdf
  python pdf_to_xlsx.py report.pdf --pages 1-5
  python pdf_to_xlsx.py report.pdf --outfile tables.xlsx
  python pdf_to_xlsx.py report.pdf --pages all --outfile output/tables.xlsx

示例(批量模式):
  python pdf_to_xlsx.py --batch ./invoices/
  python pdf_to_xlsx.py --batch ./invoices/ --outfile batch_result.xlsx
        """,
    )

    # 互斥组:单文件 input 与 --batch 不能同时指定
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "input", nargs="?", default=None,
        help="输入 PDF 文件路径(单文件模式)",
    )
    group.add_argument(
        "--batch", metavar="FOLDER",
        help="批量模式:扫描指定目录下所有 .pdf 文件,合并输出到同一 Excel",
    )

    parser.add_argument(
        "--pages", default="all",
        help="页面范围,如 'all'、'1-3'、'1,3,5',默认 all(单文件模式有效)",
    )
    parser.add_argument(
        "--outfile", default=None,
        help="输出 Excel 文件路径(单文件默认同名同目录;批量模式默认 ./pdf_batch_tables.xlsx)",
    )

    args = parser.parse_args()

    if args.batch:
        batch_pdf_to_xlsx(folder=args.batch, outfile=args.outfile)
    else:
        check_scanned_and_warn(args.input)
        pdf_to_xlsx(
            pdf_path=args.input,
            pages_str=args.pages,
            outfile=args.outfile,
        )


if __name__ == "__main__":
    main()
